import pandas as pd
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import time


start_time = time.time()
file_name = "123456"       # source of data
file1_name = "654321"      # ready empty table
df = pd.read_excel(f'/Users/User/Desktop/{file_name}.xlsx')    # import file into pandas dataframe
lst = ['boşdayanma', 'i̇stismara yararsız', 'məzuniyyət', 'sürücüsüz', 'xəstə']  # items for filtering dataframe column
values = []
workbook = load_workbook(f'/Users/User/Desktop/{file1_name}.xlsx')    # import ready empty table
wb = workbook.active
# black color to fill need cells
blackfill = PatternFill(fill_type='solid', start_color='00000000', end_color='00000000')
current_date = date.today().strftime("%d.%m.%Y")
# insert header with current date to ready empty table
wb['B1'] = f'{current_date}-ci il tarixində 2 saylı TND-nin tərkibində olan nəqliyyat vasitələri haqqında məlumat'

df.columns = df.iloc[0]    # convert first row to column header
df.drop(df.index[0], inplace=True)   # delete first row
df = df[df['Qeyd'] != 0]    # change dataframe deleting the rows in column that contains 0 (zero) values
df.dropna(subset=['Qeyd'], inplace=True)    # change dataframe deleting the rows in column that contains NULL values
df = df.iloc[:, 0:17]    # to keep 17 columns
df.drop(df.iloc[:, 5:13], inplace=True, axis=1)         # delete columns
df.drop(df.iloc[:, [0, 5, 7]], inplace=True, axis=1)    # delete columns
df.rename({1.0: 'Sahə'}, axis=1, inplace=True)          # rename column name
df['Sahə'] = df['Sahə'].astype(int)                     # convert float type to integer
df['Qeyd'] = df['Qeyd'].str.lower()                     # convert column values to lowercase
df = df[df['Qeyd'].isin(lst)]                           # filter a column by a list
df.sort_values(by='Qeyd', ascending=True, inplace=True)  # sort column by ascending


def create_table():
    rows_count = []

    try:
        rows_count.append(df['Qeyd'].value_counts()[lst[0]])  # count of boşdayanma
    except:
        pass

    try:
        rows_count.append(df['Qeyd'].value_counts()[lst[1]])  # count oḟ istismara yararsız
    except:
        pass

    try:
        rows_count.append(df['Qeyd'].value_counts()[lst[2]])  # count of məzuniyyət
    except:
        pass

    try:
        rows_count.append(df['Qeyd'].value_counts()[lst[3]])  # count of sürücüsüz
    except:
        pass

    try:
        rows_count.append(df['Qeyd'].value_counts()[lst[4]])  # count of xəstə
    except:
        pass

    # iterate rows by cell and add values to list
    for row in range(0, len(df)):
        for col in range(0, len(df.columns)):
            values.append(df.iloc[row:row+1, col].tolist()[0])

    # iterate rows by cell and fill need cells with color
    i = 0  # it will be need for indexation of 'rows_count' list
    e = 3
    for r in range(e+1, int(len(values)/6)+int(len(rows_count)-1)*2):
        if r == e + 1 + rows_count[i]:
            for col in range(1, 8):
                wb.cell(row=r, column=col).fill = blackfill
                wb.cell(row=r, column=1).value = '.'
                wb.cell(row=r, column=2).value = '.'
            i += 1
            e = r

        else:
            for col in range(2, 8):
                wb.cell(row=r, column=col).value = values[0]
                values.pop(0)

    # delete empty rows
    empty_row = 0
    for empty_cell in wb['B']:
        empty_row += 1
        if empty_cell.value is not None:
            pass
        else:
            wb.delete_rows(empty_row, wb.max_row - empty_row)
            break
    workbook.save('/Users/User/Desktop/final_table.xlsx')
    workbook.close()


end_time = time.time()
print("The program's execution time: " + str(round(end_time-start_time)) + " seconds")

if __name__ == '__main__':
    create_table()
