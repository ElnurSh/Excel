# extract excel cell comment from one column (A) and insert to the other column (B) via python (openpyxl)

from openpyxl import load_workbook


workbook = load_workbook('/Users/User/Desktop/test.xlsx')
workbook1 = workbook.active
first_sheet = workbook.get_sheet_names()[0]
worksheet = workbook.get_sheet_by_name(first_sheet)

lst = []
num = 0

for data in worksheet['A']:
    lst.append(data.comment.text)
    num += 1
    workbook1.cell(row=num, column=2).value = lst[0]
    lst.pop(0)

workbook.save('/Users/User/Desktop/test.xlsx')
